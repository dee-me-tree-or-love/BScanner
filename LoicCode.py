'''
from openpyxl import load_workbook
wb = load_workbook(filename='students.xlsx', read_only=True)
ws = wb['alle'] # ws is now an IterableWorksheet

def getValue(studentNumber):
	for row in ws.rows:
		for cell in row:
			if(studentNumber == cell.value):
				return row
			#print(cell.value)

while (True):

	print "Enter a Student Number:",
	studentNr = raw_input()
	print repr(studentNr)
	print studentNr

	student = getValue(int(studentNr))


	for cell in student:
		print(cell.column)
		print(cell.value)
'''
from openpyxl import load_workbook
wb = load_workbook(filename='students.xlsx', read_only=True)
ws = wb['alle'] # ws is now an IterableWorksheet

def getValue(studentNumber):
	for row in ws.rows:
		for cell in row:
			if(studentNumber == cell.value):
				return row
			#print(cell.value)

while (True):

	print ("Enter a Student Number:")
	studentNr = input()
	print (repr(studentNr))
	print (studentNr)

	student = getValue(int(studentNr))


	for cell in student:
		print(cell.column)
		print(cell.value)
