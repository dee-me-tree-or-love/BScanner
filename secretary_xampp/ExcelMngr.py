from openpyxl import load_workbook

class ExcelManager:
    # initialised with a workbook
    wb = load_workbook(filename='students.xlsx', read_only=True)
    ws = wb['alle']

    def __init__(self):
        wb = load_workbook(filename='students.xlsx', read_only=True)
        ws = wb['alle']  # ws is now an IterableWorksheet

    # to find if a student with this number exists
    # returns True, if yes - False, if not
    def testStudentNumber(self,studentNumber):
        result = False
        for row in self.ws.rows:

            for cell in row:
                if (studentNumber == cell.value):

                    result = True
                    return result
                    # print(cell.value)
                    # should be done now...
        return result

    # to retrieve all student data by number
    # returns null if no such student exists
    def getStudentInfo(self,studentNumber):
        for row in self.ws.rows:
            for cell in row:
                if (studentNumber == cell.value):
                    return row
                    # print(cell.value)
        return None
    '''TODO!
                #a method to write shit (an array with numbers) to json file
    '''

#print("shit")

''' Not necessary right now
while (True):

    print("Enter a Student Number:")
    studentNr = input()
    print(repr(studentNr))
    print(studentNr)

    student = getValue(int(studentNr))

    for cell in student:
        print(cell.column)
        print(cell.value)
'''
